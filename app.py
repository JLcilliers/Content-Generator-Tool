"""
SEO Content Brief Generator - Streamlit Application
Main web interface for generating AI-powered content briefs.
"""

import streamlit as st
import os
import zipfile
import io
from datetime import datetime

# Excel processing
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Load environment variables
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Import application modules
from ai_provider import AIProvider
from web_researcher import WebResearcher
from brief_generator import BriefGenerator
from document_formatter import DocumentFormatter, generate_markdown_brief
from validators import validate_brief, validate_page_title, validate_meta_description
from client_guidelines import (
    get_client_guidelines,
    get_client_name_from_url,
    is_known_client,
    get_language_preference
)

# Try to import Supabase manager
SUPABASE_AVAILABLE = False
try:
    from supabase_client_manager import SupabaseClientManager
    SUPABASE_AVAILABLE = True
except Exception:
    pass


# Page configuration
st.set_page_config(
    page_title="SEO Content Brief Generator",
    page_icon="📝",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1E3A5F;
        margin-bottom: 1rem;
    }
    .section-header {
        font-size: 1.2rem;
        font-weight: 600;
        color: #2C5282;
        margin-top: 1rem;
        margin-bottom: 0.5rem;
    }
    .success-box {
        padding: 1rem;
        background-color: #C6F6D5;
        border-radius: 0.5rem;
        border-left: 4px solid #38A169;
    }
    .warning-box {
        padding: 1rem;
        background-color: #FEFCBF;
        border-radius: 0.5rem;
        border-left: 4px solid #D69E2E;
    }
    .error-box {
        padding: 1rem;
        background-color: #FED7D7;
        border-radius: 0.5rem;
        border-left: 4px solid #E53E3E;
    }
    .info-box {
        padding: 1rem;
        background-color: #BEE3F8;
        border-radius: 0.5rem;
        border-left: 4px solid #3182CE;
    }
    .stButton>button {
        width: 100%;
    }
</style>
""", unsafe_allow_html=True)


def init_session_state():
    """Initialize session state variables."""
    # Single brief mode
    if 'brief_data' not in st.session_state:
        st.session_state.brief_data = None
    if 'research_data' not in st.session_state:
        st.session_state.research_data = None
    if 'internal_links' not in st.session_state:
        st.session_state.internal_links = []
    if 'generation_complete' not in st.session_state:
        st.session_state.generation_complete = False
    if 'document_path' not in st.session_state:
        st.session_state.document_path = None

    # Batch mode
    if 'batch_items' not in st.session_state:
        st.session_state.batch_items = []
    if 'batch_results' not in st.session_state:
        st.session_state.batch_results = []
    if 'batch_documents' not in st.session_state:
        st.session_state.batch_documents = []
    if 'batch_complete' not in st.session_state:
        st.session_state.batch_complete = False
    if 'batch_error' not in st.session_state:
        st.session_state.batch_error = None
    if 'batch_zip_data' not in st.session_state:
        st.session_state.batch_zip_data = None


def get_available_providers():
    """Get list of available AI providers."""
    return AIProvider.list_available_providers()


def render_sidebar():
    """Render the sidebar with settings and client management."""
    with st.sidebar:
        st.markdown("### Settings")

        # AI Provider selection
        available_providers = get_available_providers()

        if not available_providers:
            st.error("No AI providers configured. Please add API keys to .env file or Streamlit secrets.")
            st.stop()

        provider_names = {
            'openai': 'OpenAI GPT-5.2',
            'claude': 'Claude Opus 4.5',
            'grok': 'Grok 4',
            'perplexity': 'Perplexity Sonar Pro',
            'mistral': 'Mistral Large 3'
        }

        provider_options = [provider_names.get(p, p) for p in available_providers]
        selected_idx = st.selectbox(
            "AI Provider",
            range(len(available_providers)),
            format_func=lambda x: provider_options[x],
            help="Select which AI provider to use for generating briefs"
        )
        selected_provider = available_providers[selected_idx]

        st.markdown("---")

        # Client Management Section
        st.markdown("### Client Management")

        # Check if Supabase is available
        if SUPABASE_AVAILABLE:
            try:
                client_manager = SupabaseClientManager()
                clients = client_manager.list_clients()

                if clients:
                    selected_client = st.selectbox(
                        "Load Client Profile",
                        ["-- Select Client --"] + clients,
                        help="Load a saved client profile"
                    )

                    if selected_client != "-- Select Client --":
                        client_data = client_manager.get_client(selected_client)
                        if client_data:
                            st.success(f"Loaded: {selected_client}")
                            return selected_provider, client_data

                # Create new client
                with st.expander("Create New Client"):
                    new_client_name = st.text_input("Client Name", key="new_client_name")
                    new_client_site = st.text_input("Website URL", key="new_client_site")
                    new_client_industry = st.text_input("Industry", key="new_client_industry")

                    if st.button("Save Client"):
                        if new_client_name and new_client_site:
                            success = client_manager.create_client(
                                new_client_name,
                                {"site": new_client_site, "industry": new_client_industry}
                            )
                            if success:
                                st.success(f"Created client: {new_client_name}")
                                st.rerun()
                            else:
                                st.error("Failed to create client")
                        else:
                            st.warning("Please enter client name and website URL")

            except Exception as e:
                st.warning(f"Supabase not configured: {str(e)[:50]}...")

        else:
            st.info("Connect Supabase to save client profiles")

        st.markdown("---")

        # Known clients info
        st.markdown("### Known Clients")
        st.markdown("""
        Auto-detected clients with special rules:
        - **HealthWorks** (healthworkstx.com)
        - **CellGate** (cell-gate.com)
        - **AIM Companies** (aim-companies.com)
        """)

        return selected_provider, None


def render_main_form(selected_provider: str, client_data: dict = None):
    """Render the main brief generation form."""
    st.markdown('<p class="main-header">📝 SEO Content Brief Generator</p>', unsafe_allow_html=True)

    st.markdown("""
    Generate professional, AI-powered SEO content briefs with automatic web research,
    URL validation, and Word document export.
    """)

    # Input form
    col1, col2 = st.columns([2, 1])

    with col1:
        st.markdown('<p class="section-header">Brief Details</p>', unsafe_allow_html=True)

        # Pre-fill from client data if available
        default_url = client_data.get('site', '') if client_data else ''

        url = st.text_input(
            "Website URL",
            value=default_url,
            placeholder="https://example.com",
            help="Enter the client's website URL"
        )

        topic = st.text_input(
            "Topic",
            placeholder="e.g., Neck Pain Treatment",
            help="What is this content about?"
        )

        keywords = st.text_area(
            "Keywords (one per line, primary first)",
            placeholder="neck pain chiropractor plano\nupper cervical care for neck pain\nneck pain relief plano",
            height=120,
            help="Enter keywords with the primary keyword on the first line"
        )

    with col2:
        st.markdown('<p class="section-header">Options</p>', unsafe_allow_html=True)

        # Check for known client
        if url:
            if is_known_client(url):
                client_name = get_client_name_from_url(url)
                language = get_language_preference(url)
                st.markdown(f"""
                <div class="info-box">
                    <strong>Known Client Detected:</strong> {client_name}<br>
                    <strong>Language:</strong> {'UK English' if language == 'UK' else 'US English'}<br>
                    Special guidelines will be applied automatically.
                </div>
                """, unsafe_allow_html=True)

        auto_research = st.checkbox(
            "Auto-research website",
            value=True,
            help="Automatically analyze the website for brand voice and content"
        )

        auto_links = st.checkbox(
            "Auto-discover internal links",
            value=True,
            help="Automatically find relevant internal links"
        )

        if not auto_links:
            manual_links = st.text_area(
                "Manual Internal Links (one per line)",
                placeholder="https://example.com/about\nhttps://example.com/services\nhttps://example.com/contact",
                height=100
            )

    st.markdown("---")

    # Generate button
    generate_col1, generate_col2, generate_col3 = st.columns([1, 2, 1])

    with generate_col2:
        generate_button = st.button(
            "🚀 Generate Content Brief",
            type="primary",
            use_container_width=True
        )

    if generate_button:
        # Validate inputs
        if not url:
            st.error("Please enter a website URL")
            return
        if not topic:
            st.error("Please enter a topic")
            return
        if not keywords.strip():
            st.error("Please enter at least one keyword")
            return

        # Parse keywords
        keyword_list = [k.strip() for k in keywords.strip().split('\n') if k.strip()]
        if len(keyword_list) < 1:
            st.error("Please enter at least one keyword")
            return

        primary_keyword = keyword_list[0]
        secondary_keywords = keyword_list[1:] if len(keyword_list) > 1 else []

        # Generate brief
        with st.spinner("Generating content brief..."):
            try:
                progress = st.progress(0, text="Initializing...")

                # Step 1: Web research
                research_data = None
                if auto_research:
                    progress.progress(10, text="Researching website...")
                    researcher = WebResearcher()
                    research_data = researcher.research_website(url, topic)
                    st.session_state.research_data = research_data

                # Step 2: Find internal links
                progress.progress(30, text="Finding internal links...")
                if auto_links:
                    researcher = WebResearcher()
                    internal_links = researcher.find_internal_links(url, topic, keyword_list)
                else:
                    manual_link_list = [l.strip() for l in manual_links.strip().split('\n') if l.strip()]
                    internal_links = manual_link_list[:3]

                # Ensure we have 3 links
                if len(internal_links) < 3:
                    st.warning(f"Only found {len(internal_links)} internal links. Brief may be incomplete.")

                st.session_state.internal_links = internal_links

                # Step 3: Generate brief
                progress.progress(50, text="Generating brief with AI...")
                generator = BriefGenerator(provider=selected_provider)
                brief_data = generator.generate_brief(
                    url=url,
                    topic=topic,
                    primary_keyword=primary_keyword,
                    secondary_keywords=secondary_keywords,
                    internal_links=internal_links,
                    website_research=research_data
                )

                st.session_state.brief_data = brief_data

                # Step 4: Create document
                progress.progress(80, text="Creating Word document...")
                formatter = DocumentFormatter()
                doc_path = formatter.create_brief_document(brief_data)
                st.session_state.document_path = doc_path

                progress.progress(100, text="Complete!")
                st.session_state.generation_complete = True

            except Exception as e:
                st.error(f"Error generating brief: {str(e)}")
                import traceback
                st.expander("Error Details").code(traceback.format_exc())


def render_results():
    """Render the generated brief results."""
    if not st.session_state.generation_complete or not st.session_state.brief_data:
        return

    brief_data = st.session_state.brief_data

    st.markdown("---")
    st.markdown('<p class="main-header">📄 Generated Content Brief</p>', unsafe_allow_html=True)

    # Download button
    if st.session_state.document_path and os.path.exists(st.session_state.document_path):
        with open(st.session_state.document_path, 'rb') as f:
            st.download_button(
                label="📥 Download Word Document",
                data=f.read(),
                file_name=os.path.basename(st.session_state.document_path),
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                type="primary"
            )

    # Validation results
    validation = brief_data.get('_validation', {})
    if validation.get('errors'):
        st.markdown('<div class="error-box">', unsafe_allow_html=True)
        st.markdown("**Validation Errors:**")
        for error in validation['errors']:
            st.markdown(f"- {error}")
        st.markdown('</div>', unsafe_allow_html=True)
    elif validation.get('warnings'):
        st.markdown('<div class="warning-box">', unsafe_allow_html=True)
        st.markdown("**Validation Warnings:**")
        for warning in validation['warnings']:
            st.markdown(f"- {warning}")
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="success-box">', unsafe_allow_html=True)
        st.markdown("**All validations passed!**")
        st.markdown('</div>', unsafe_allow_html=True)

    # Brief preview
    st.markdown("### Brief Preview")

    # Create tabs for different sections
    tab1, tab2, tab3, tab4 = st.tabs(["Overview", "Structure", "Guidelines", "Headings & FAQs"])

    with tab1:
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Client:**")
            st.write(brief_data.get('client_name', ''))
            st.markdown("**Site:**")
            st.write(brief_data.get('site', ''))
            st.markdown("**Topic:**")
            st.write(brief_data.get('topic', ''))
        with col2:
            st.markdown("**Primary Keyword:**")
            st.write(brief_data.get('primary_keyword', ''))
            st.markdown("**Secondary Keywords:**")
            st.write(', '.join(brief_data.get('secondary_keywords', [])))

    with tab2:
        st.markdown("**Page Type:**")
        st.write(brief_data.get('page_type', ''))

        st.markdown("**Page Title:**")
        title = brief_data.get('page_title', '')
        title_valid, title_msg = validate_page_title(title)
        st.write(f"{title} ({len(title)} chars)")
        if not title_valid:
            st.warning(title_msg)

        st.markdown("**Meta Description:**")
        meta = brief_data.get('meta_description', '')
        meta_valid, meta_msg = validate_meta_description(meta)
        st.write(f"{meta} ({len(meta)} chars)")
        if not meta_valid:
            st.warning(meta_msg)

        st.markdown("**Target URL:**")
        st.write(brief_data.get('target_url', ''))

        st.markdown("**H1 Heading:**")
        st.write(brief_data.get('h1', ''))

        st.markdown("**Internal Links:**")
        for link in brief_data.get('internal_links', []):
            st.write(f"- {link}")

    with tab3:
        st.markdown("**Word Count:**")
        st.write(brief_data.get('word_count', ''))

        st.markdown("**Audience:**")
        for item in brief_data.get('audience', []):
            st.write(f"- {item}")

        st.markdown("**Tone:**")
        for item in brief_data.get('tone', []):
            st.write(f"- {item}")

        st.markdown("**CTA:**")
        st.write(brief_data.get('cta', ''))

        st.markdown("**Restrictions:**")
        for item in brief_data.get('restrictions', []):
            st.write(f"- {item}")

    with tab4:
        st.markdown("**Heading Structure:**")
        for heading in brief_data.get('headings', []):
            level = heading.get('level', '')
            text = heading.get('text', '')
            desc = heading.get('description', '')
            st.markdown(f"**{level} - {text}**")
            if desc:
                st.markdown(f"_{desc}_")

            for sub in heading.get('subheadings', []):
                st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;**H3 - {sub.get('text', '')}**")
                if sub.get('description'):
                    st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;_{sub.get('description')}_")

        st.markdown("---")
        st.markdown("**FAQs:**")
        for faq in brief_data.get('faqs', []):
            st.write(f"- {faq}")

    # Full markdown preview
    with st.expander("View Full Markdown"):
        markdown_brief = generate_markdown_brief(brief_data)
        st.code(markdown_brief, language="markdown")


def parse_excel_file(uploaded_file):
    """Parse uploaded Excel file and extract brief items.

    Expected format:
    Column A: Website URL
    Column B: Topic
    Column C: Primary Keyword
    Column D: Secondary Keyword 1
    Column E: Secondary Keyword 2
    Column F: Secondary Keyword 3
    """
    items = []

    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active

        # Skip header row, process data rows
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or not row[0]:
                continue

            url = str(row[0]).strip() if row[0] else ''
            topic = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            primary_kw = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            secondary_kw1 = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            secondary_kw2 = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            secondary_kw3 = str(row[5]).strip() if len(row) > 5 and row[5] else ''

            # Validate required fields
            if not url or not topic or not primary_kw:
                continue

            # Build secondary keywords list (filter out empty)
            secondary_keywords = [kw for kw in [secondary_kw1, secondary_kw2, secondary_kw3] if kw]

            items.append({
                'row': row_num,
                'url': url,
                'topic': topic,
                'primary_keyword': primary_kw,
                'secondary_keywords': secondary_keywords
            })

        workbook.close()

    except Exception as e:
        raise ValueError(f"Failed to parse Excel file: {str(e)}")

    return items


def process_batch(items: list, selected_provider: str, progress_bar, status_text):
    """Process batch of brief items and generate documents.

    Returns:
        tuple: (documents list, error message or None)
    """
    documents = []
    results = []

    total = len(items)

    for idx, item in enumerate(items):
        current = idx + 1
        progress = int((current / total) * 100)

        status_text.text(f"Processing {current}/{total}: {item['topic']}")
        progress_bar.progress(progress)

        try:
            # Initialize components
            researcher = WebResearcher()
            generator = BriefGenerator(provider=selected_provider)
            formatter = DocumentFormatter()

            # Step 1: Web research
            research_data = researcher.research_website(item['url'], item['topic'])

            # Step 2: Find internal links
            all_keywords = [item['primary_keyword']] + item['secondary_keywords']
            internal_links = researcher.find_internal_links(item['url'], item['topic'], all_keywords)

            # Step 3: Generate brief
            brief_data = generator.generate_brief(
                url=item['url'],
                topic=item['topic'],
                primary_keyword=item['primary_keyword'],
                secondary_keywords=item['secondary_keywords'],
                internal_links=internal_links,
                website_research=research_data
            )

            # Step 4: Create document
            doc_path = formatter.create_brief_document(brief_data)

            documents.append(doc_path)
            results.append({
                'row': item['row'],
                'topic': item['topic'],
                'status': 'success',
                'path': doc_path
            })

        except Exception as e:
            error_msg = f"Error on row {item['row']} ({item['topic']}): {str(e)}"
            return documents, results, error_msg

    return documents, results, None


def create_zip_archive(document_paths: list) -> bytes:
    """Create a ZIP archive containing all generated documents."""
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for doc_path in document_paths:
            if os.path.exists(doc_path):
                # Use just the filename in the ZIP
                zip_file.write(doc_path, os.path.basename(doc_path))

    zip_buffer.seek(0)
    return zip_buffer.getvalue()


def render_batch_form(selected_provider: str):
    """Render the batch upload form for processing multiple briefs."""
    st.markdown('<p class="main-header">📊 Batch Brief Generator</p>', unsafe_allow_html=True)

    st.markdown("""
    Upload an Excel file to generate multiple content briefs at once.
    All briefs will be bundled into a ZIP file for download.
    """)

    # Check if Excel processing is available
    if not EXCEL_AVAILABLE:
        st.error("Excel processing not available. Please install openpyxl: `pip install openpyxl`")
        return

    # File format instructions
    with st.expander("Excel File Format", expanded=True):
        st.markdown("""
        Your Excel file should have the following columns:

        | Column A | Column B | Column C | Column D | Column E | Column F |
        |----------|----------|----------|----------|----------|----------|
        | Website URL | Topic | Primary Keyword | Secondary KW 1 | Secondary KW 2 | Secondary KW 3 |

        **Notes:**
        - First row should be headers (will be skipped)
        - Columns A, B, C are required
        - Columns D, E, F are optional secondary keywords
        - Empty rows will be skipped
        """)

    # File uploader
    uploaded_file = st.file_uploader(
        "Upload Excel File",
        type=['xlsx'],
        help="Upload an .xlsx file with your brief data"
    )

    if uploaded_file:
        # Parse the file
        try:
            items = parse_excel_file(uploaded_file)
            st.session_state.batch_items = items

            if not items:
                st.warning("No valid rows found in the Excel file. Please check the format.")
                return

            st.success(f"Found {len(items)} briefs to generate")

            # Preview table
            st.markdown("### Preview")
            preview_data = []
            for item in items:
                preview_data.append({
                    'Row': item['row'],
                    'URL': item['url'][:50] + '...' if len(item['url']) > 50 else item['url'],
                    'Topic': item['topic'],
                    'Primary KW': item['primary_keyword'],
                    'Secondary KWs': ', '.join(item['secondary_keywords'])
                })
            st.dataframe(preview_data, use_container_width=True)

        except ValueError as e:
            st.error(str(e))
            return

    st.markdown("---")

    # Generate button
    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        generate_batch = st.button(
            "🚀 Generate All Briefs",
            type="primary",
            use_container_width=True,
            disabled=not st.session_state.batch_items
        )

    if generate_batch and st.session_state.batch_items:
        # Reset batch state
        st.session_state.batch_complete = False
        st.session_state.batch_error = None
        st.session_state.batch_documents = []
        st.session_state.batch_results = []
        st.session_state.batch_zip_data = None

        # Process batch
        progress_bar = st.progress(0)
        status_text = st.empty()

        with st.spinner("Processing batch..."):
            documents, results, error = process_batch(
                st.session_state.batch_items,
                selected_provider,
                progress_bar,
                status_text
            )

            st.session_state.batch_documents = documents
            st.session_state.batch_results = results

            if error:
                st.session_state.batch_error = error
                progress_bar.empty()
                status_text.empty()
                st.error(f"Batch processing stopped: {error}")
            else:
                # Create ZIP archive
                if documents:
                    st.session_state.batch_zip_data = create_zip_archive(documents)
                    st.session_state.batch_complete = True

                progress_bar.progress(100)
                status_text.text("Complete!")

    # Show results
    render_batch_results()


def render_batch_results():
    """Render batch processing results."""
    if not st.session_state.batch_results:
        return

    st.markdown("---")
    st.markdown("### Results")

    # Summary
    total = len(st.session_state.batch_items)
    completed = len(st.session_state.batch_results)

    if st.session_state.batch_error:
        st.markdown(f"""
        <div class="warning-box">
            <strong>Partial Completion:</strong> {completed}/{total} briefs generated before error
        </div>
        """, unsafe_allow_html=True)
    elif st.session_state.batch_complete:
        st.markdown(f"""
        <div class="success-box">
            <strong>Success!</strong> All {completed} briefs generated successfully
        </div>
        """, unsafe_allow_html=True)

    # Results table
    results_data = []
    for result in st.session_state.batch_results:
        results_data.append({
            'Row': result['row'],
            'Topic': result['topic'],
            'Status': '✅ Success' if result['status'] == 'success' else '❌ Failed'
        })

    if results_data:
        st.dataframe(results_data, use_container_width=True)

    # Download ZIP button
    if st.session_state.batch_zip_data:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="📥 Download All Briefs (ZIP)",
            data=st.session_state.batch_zip_data,
            file_name=f"content_briefs_{timestamp}.zip",
            mime="application/zip",
            type="primary"
        )


def main():
    """Main application entry point."""
    init_session_state()

    # Render sidebar and get settings
    selected_provider, client_data = render_sidebar()

    # Create tabs for single vs batch mode
    tab_single, tab_batch = st.tabs(["📝 Single Brief", "📊 Batch Upload"])

    with tab_single:
        # Render main form
        render_main_form(selected_provider, client_data)

        # Render results if available
        render_results()

    with tab_batch:
        # Render batch upload form
        render_batch_form(selected_provider)

    # Footer
    st.markdown("---")
    st.markdown(
        "<div style='text-align: center; color: #666;'>"
        "SEO Content Brief Generator | Built for Ghost Rank"
        "</div>",
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()
