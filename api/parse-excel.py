"""
Parse Excel API - Parses uploaded Excel files for batch processing
"""
import json
import sys
import os
import io

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'lib'))

from http.server import BaseHTTPRequestHandler


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            from openpyxl import load_workbook

            # Get content type and boundary for multipart parsing
            content_type = self.headers.get('Content-Type', '')
            content_length = int(self.headers.get('Content-Length', 0))

            # Read the raw body
            body = self.rfile.read(content_length)

            # Simple multipart parser to extract file
            if 'multipart/form-data' in content_type:
                # Extract boundary
                boundary = content_type.split('boundary=')[1].encode()

                # Split by boundary
                parts = body.split(b'--' + boundary)

                file_content = None
                for part in parts:
                    if b'filename=' in part:
                        # Find the file content (after double newline)
                        header_end = part.find(b'\r\n\r\n')
                        if header_end != -1:
                            file_content = part[header_end + 4:]
                            # Remove trailing boundary markers
                            if file_content.endswith(b'\r\n'):
                                file_content = file_content[:-2]
                            if file_content.endswith(b'--'):
                                file_content = file_content[:-2]
                            if file_content.endswith(b'\r\n'):
                                file_content = file_content[:-2]
                            break

                if not file_content:
                    raise ValueError("No file found in request")

                # Parse Excel
                workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
                sheet = workbook.active

                items = []
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not row[0]:
                        continue

                    url = str(row[0]).strip() if row[0] else ''
                    topic = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    primary_kw = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                    secondary_kw1 = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                    secondary_kw2 = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                    secondary_kw3 = str(row[5]).strip() if len(row) > 5 and row[5] else ''

                    if not url or not topic or not primary_kw:
                        continue

                    secondary_keywords = [kw for kw in [secondary_kw1, secondary_kw2, secondary_kw3] if kw]

                    items.append({
                        'row': row_num,
                        'url': url,
                        'topic': topic,
                        'primaryKeyword': primary_kw,
                        'secondaryKeywords': secondary_keywords
                    })

                workbook.close()

                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({'items': items}).encode())

            else:
                raise ValueError("Expected multipart/form-data")

        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
